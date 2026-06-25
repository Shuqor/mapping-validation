import os
import json
import pathlib
from openpyxl import Workbook
import core.validate as validate_module

class DummyMonkeypatch:
    def setattr(self, target, name, value):
        setattr(target, name, value)

def _write_xml(path, body: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write(body)

def _patch_rules(monkeypatch, rules):
    monkeypatch.setattr(validate_module, "read_mapping_table", lambda *args, **kwargs: None)
    monkeypatch.setattr(validate_module, "extract_rules", lambda df: rules)
    monkeypatch.setattr(
        validate_module,
        "get_parser_diagnostics",
        lambda df: {
            "status": "clean",
            "confidence": "high",
            "warnings": [],
            "info": [],
            "sheet_name": "Mapping",
            "header_row": 1,
            "rule_count": len(rules),
        },
    )

def run_repro():
    tmp_path = pathlib.Path("repro_tmp")
    tmp_path.mkdir(exist_ok=True)
    
    spec_xlsx = tmp_path / "lookup_hint_lock_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Timezone Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "UTC"

    ws["A6"] = "Country Code Lookup"
    ws["A7"] = "Source Code"
    ws["B7"] = "Target Name"
    ws["A8"] = "US"
    ws["B8"] = "United States"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>US</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>United States</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Check LookUp-Conversion Tab(Country Code) and map",
            "note": "",
        }
    ]
    
    monkeypatch = DummyMonkeypatch()
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    print(f"Summary Status: {result['summary']['status']}")
    print(f"Errors: {json.dumps(result['errors'], indent=2)}")
    print(f"Grouped Error Counts: {json.dumps(result['summary']['grouped_error_counts'], indent=2)}")
    print(f"Warnings: {json.dumps(result['warnings'], indent=2)}")
    
    if result['rule_decisions']:
        print(f"First Rule Decision: {json.dumps(result['rule_decisions'][0], indent=2)}")
    else:
        print("No rule decisions found.")

if __name__ == "__main__":
    run_repro()

import core.validate as validate_module
from openpyxl import Workbook


def _write_xml(path, body: str) -> None:
    path.write_text(body, encoding="utf-8")


def _patch_rules(monkeypatch, rules):
    monkeypatch.setattr(validate_module, "read_mapping_table", lambda *_args, **_kwargs: object())
    monkeypatch.setattr(validate_module, "extract_rules", lambda _df: rules)
    monkeypatch.setattr(
        validate_module,
        "get_parser_diagnostics",
        lambda _df: {
            "status": "clean",
            "confidence": "high",
            "warnings": [],
            "info": [],
            "sheet_name": "Mapping",
            "header_row": 1,
            "rule_count": len(rules),
        },
    )


def _build_lookup_workbook(path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Timezone Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "EST"
    ws["B3"] = "America/New_York"
    ws["A4"] = "UTC"
    ws["B4"] = "UTC"

    ws["A6"] = "Country Code Lookup"
    ws["A7"] = "Source Code"
    ws["B7"] = "Target Name"
    ws["A8"] = "US"
    ws["B8"] = "United States"
    ws["A9"] = "AU"
    ws["B9"] = "Australia"

    workbook.create_sheet("Notes")
    workbook.save(path)


def test_lookup_table_mapping_resolves_value_from_nonstandard_tab(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_spec.xlsx"
    _build_lookup_workbook(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"

    _write_xml(
        src_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>EST</code></status>\n',
    )
    _write_xml(
        tgt_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><timezone>America/New_York</timezone></status>\n',
    )

    rules = [
        {
            "target_xpath": "/status/timezone",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "If source use the timezone lookup table then map the value to target",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["status"] == "PASS"
    assert result["error_count"] == 0
    assert result["rule_support_summary"]["lookup_table_rules"] == 1
    decision = result["rule_decisions"][0]
    lookup_resolution = decision.get("lookup_resolution", {})
    assert lookup_resolution.get("status") == "found"
    assert lookup_resolution.get("trace", {}).get("binding_mode") in {"hint_locked", "hint_preferred", "score_fallback"}


def test_lookup_table_rule_falls_back_to_source_when_configured(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_spec.xlsx"
    _build_lookup_workbook(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"

    _write_xml(
        src_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>UNKNOWN_CODE</code></status>\n',
    )
    _write_xml(
        tgt_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>UNKNOWN_CODE</country></status>\n',
    )

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Check LookUp-Conversion Tab(Country Code) and map. If cannot find in LookUp-Conversion then map the source",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["status"] == "PASS"
    assert result["error_count"] == 0
    assert result["rule_support_summary"]["lookup_table_rules"] == 1


def test_lookup_table_missing_key_without_fallback_reports_mismatch(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_spec.xlsx"
    _build_lookup_workbook(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"

    _write_xml(
        src_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>UNKNOWN_CODE</code></status>\n',
    )
    _write_xml(
        tgt_xml,
        '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>UNKNOWN_CODE</country></status>\n',
    )

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Refer lookup country code",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) == 1
    assert any("Lookup key is not present" in err for err in result["errors"])


def test_lookup_table_duplicate_key_conflict_is_conservative(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_conflict_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Country Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "United States"
    ws["A4"] = "US"
    ws["B4"] = "USA"
    ws["A5"] = "AU"
    ws["B5"] = "Australia"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>US</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>US</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Refer lookup country code",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) == 0
    assert any("lookup conflict finding(s) were downgraded" in warning for warning in result["warnings"])


def test_lookup_table_hint_lock_prefers_matching_lookup_block(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_hint_lock_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Timezone Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "UTC"
    ws["A4"] = "AU"
    ws["B4"] = "AEST"

    ws["A6"] = "Country Code Lookup"
    ws["A7"] = "Source Code"
    ws["B7"] = "Target Name"
    ws["A8"] = "US"
    ws["B8"] = "United States"
    ws["A9"] = "AU"
    ws["B9"] = "Australia"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>US</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>United States</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Check LookUp-Conversion Tab(Country Code) and map",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["status"] == "PASS"
    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) == 0
    lookup_resolution = result["rule_decisions"][0].get("lookup_resolution", {})
    assert lookup_resolution.get("status") in {"found", "ambiguous"}
    trace = lookup_resolution.get("trace", {})
    top_candidates = trace.get("top_candidates", [])
    assert top_candidates
    assert top_candidates[0].get("sheet") == "RandomRefData"


def test_lookup_table_composite_key_resolution(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_composite_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "CompositeRef"

    ws["A1"] = "Country Composite Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Qualifier"
    ws["C2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "SEA"
    ws["C3"] = "United States Sea"
    ws["A4"] = "US"
    ws["B4"] = "AIR"
    ws["C4"] = "United States Air"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>US|SEA</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>United States Sea</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Refer lookup country code",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["status"] == "PASS"
    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) == 0


def test_lookup_table_tolerates_single_spacer_row_in_block(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_spacer_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Country Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "United States"
    ws["A4"] = ""
    ws["B4"] = ""
    ws["A5"] = "AU"
    ws["B5"] = "Australia"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>AU</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>Australia</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Refer lookup country code",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(str(spec_xlsx), str(src_xml), str(tgt_xml))

    assert result["summary"]["status"] == "PASS"
    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) == 0


def test_lookup_strict_fails_low_confidence_conflict(tmp_path, monkeypatch):
    spec_xlsx = tmp_path / "lookup_conflict_strict_spec.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "RandomRefData"

    ws["A1"] = "Country Lookup"
    ws["A2"] = "Code"
    ws["B2"] = "Mapped Value"
    ws["A3"] = "US"
    ws["B3"] = "United States"
    ws["A4"] = "US"
    ws["B4"] = "USA"
    ws["A5"] = "AU"
    ws["B5"] = "Australia"
    workbook.save(spec_xlsx)

    src_xml = tmp_path / "input.xml"
    tgt_xml = tmp_path / "output.xml"
    _write_xml(src_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><code>US</code></status>\n')
    _write_xml(tgt_xml, '<?xml version="1.0" encoding="UTF-8"?>\n<status><country>US</country></status>\n')

    rules = [
        {
            "target_xpath": "/status/country",
            "source_xpath": "/status/code",
            "cardinality": "",
            "condition": "Refer lookup country code",
            "note": "",
        }
    ]
    _patch_rules(monkeypatch, rules)

    result = validate_module.validate_mapping(
        str(spec_xlsx),
        str(src_xml),
        str(tgt_xml),
        validation_mode="lookup_strict",
    )

    assert result["summary"]["status"] == "FAIL"
    assert result["summary"]["grouped_error_counts"].get("lookup_mismatches", 0) >= 1

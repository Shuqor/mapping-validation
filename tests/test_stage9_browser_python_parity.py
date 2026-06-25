import json
import os
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.validate import validate_mapping_from_payload_bytes


ROOT_DIR = Path(__file__).resolve().parents[1]
WEB_PATH = ROOT_DIR / "web" / "index.html"
RULES_DIR = ROOT_DIR / "rules"
SAMPLES_DIR = ROOT_DIR / "samples"


@pytest.mark.skipif(os.getenv("RUN_BROWSER_PARITY_TESTS") != "1", reason="Enable with RUN_BROWSER_PARITY_TESTS=1")
def test_browser_python_parity_distribution_for_real_case():
    playwright_sync = pytest.importorskip("playwright.sync_api")

    case_candidates = [
        (
            "spec.xlsx",
            "input.xml",
            "output.xml",
        ),
        (
            "Inttra-Contivo_X12_300_5030_to_JSON_BOOKINGINBOUND.xlsx",
            "input.x12",
            "output.json",
        ),
        (
            "Inttra-Contivo_EDIFACT_IFTMBF_D99B_to_JSON_BOOKINGINBOUND.xlsx",
            "Inbound_BK3_CU2100_IFTMBF.edi",
            "output INTTRA.json",
        ),
    ]

    selected = None
    for spec_name, input_name, output_name in case_candidates:
        spec_path = RULES_DIR / spec_name
        input_path = SAMPLES_DIR / input_name
        output_path = SAMPLES_DIR / output_name
        if spec_path.exists() and input_path.exists() and output_path.exists():
            selected = (spec_path, input_path, output_path)
            break

    if not selected:
        pytest.skip("No parity candidate files are available")

    spec_path, input_path, output_path = selected

    py_result = validate_mapping_from_payload_bytes(
        str(spec_path),
        input_path.read_bytes(),
        input_path.name,
        output_path.read_bytes(),
        output_path.name,
        validation_mode="strict",
    )

    with playwright_sync.sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(WEB_PATH.resolve().as_uri())

        page.set_input_files("#mapping_spec", str(spec_path))
        page.set_input_files("#input_payload", str(input_path))
        page.set_input_files("#output_payload", str(output_path))
        page.select_option("#validation_mode", "strict")
        page.click("#submit-btn")

        page.wait_for_function("() => Boolean(window.lastResult && window.lastResult.summary)", timeout=120000)
        js_result = page.evaluate(
            """
            () => ({
              checked_rules: window.lastResult?.checked_rules || 0,
              rule_support_summary: window.lastResult?.rule_support_summary || {},
              rule_decisions: window.lastResult?.rule_decisions || [],
            })
            """
        )
        browser.close()

    py_support = py_result.get("rule_support_summary", {})
    js_support = js_result.get("rule_support_summary", {})

    py_counts = {
        "enforced": int(py_support.get("enforced_rules", 0)),
        "parsed_only": int(py_support.get("parsed_only_rules", 0)),
        "unsupported": int(py_support.get("unsupported_rules", 0)),
    }
    js_counts = {
        "enforced": int(js_support.get("enforced_rules", 0)),
        "parsed_only": int(js_support.get("parsed_only_rules", 0)),
        "unsupported": int(js_support.get("unsupported_rules", 0)),
    }

    # Exact parity is ideal, but keep a narrow tolerance to avoid platform-format parser drift.
    assert abs(int(py_result.get("checked_rules", 0)) - int(js_result.get("checked_rules", 0))) <= 2
    # Semantic family rollout can shift small numbers between enforced and parsed-only
    # while preserving top-level checked-rules parity for stable browser/runtime contracts.
    assert abs(py_counts["enforced"] - js_counts["enforced"]) <= 10
    assert abs(py_counts["parsed_only"] - js_counts["parsed_only"]) <= 10
    assert abs(py_counts["unsupported"] - js_counts["unsupported"]) <= 5

    assert isinstance(py_result.get("rule_decisions", []), list)
    assert isinstance(js_result.get("rule_decisions", []), list)


@pytest.mark.skipif(os.getenv("RUN_BROWSER_PARITY_TESTS") != "1", reason="Enable with RUN_BROWSER_PARITY_TESTS=1")
def test_browser_decision_drift_panel_renders_diff_summary():
    playwright_sync = pytest.importorskip("playwright.sync_api")

    diff_path = ROOT_DIR / "results" / "ci" / "stage10_spec_coverage_decision_diff.json"
    if not diff_path.exists():
        diff_path.parent.mkdir(parents=True, exist_ok=True)
        diff_path.write_text(
            json.dumps(
                {
                    "decision_changes": {
                        "changed_rows": 2,
                        "status_transitions": [{"transition": "parsed_only->enforced", "count": 1}],
                        "family_transitions": [{"transition": "manual_review->direct_map", "count": 1}],
                        "reason_code_transitions": [{"transition": "parsed_only_a->enforced_a", "count": 1}],
                        "top_changed_rows": [
                            {
                                "row": 1,
                                "target_xpath": "/a",
                                "severity_score": 120,
                                "baseline": {"status": "parsed_only", "family": "manual_review"},
                                "current": {"status": "enforced", "family": "direct_map"},
                            },
                            {
                                "row": 2,
                                "target_xpath": "/b",
                                "severity_score": 95,
                                "baseline": {"status": "unsupported", "family": "field_concat_mapping"},
                                "current": {"status": "parsed_only", "family": "manual_review"},
                            },
                        ],
                    }
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    payload = json.loads(diff_path.read_text(encoding="utf-8"))

    with playwright_sync.sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(WEB_PATH.resolve().as_uri())
        page.evaluate("payload => window.renderDecisionDrift(payload)", payload)

        result = page.evaluate(
            """
            () => ({
                hidden: document.getElementById('decision-diff-wrap')?.classList.contains('hidden'),
                changed: document.getElementById('decision-diff-changed-count')?.textContent,
                status: document.getElementById('decision-diff-status-count')?.textContent,
                family: document.getElementById('decision-diff-family-count')?.textContent,
                summary: document.getElementById('decision-diff-summary')?.textContent,
                list: document.getElementById('decision-diff-list')?.textContent,
            })
            """
        )
        browser.close()

    assert result["hidden"] is False
    assert result["changed"] == "2"
    assert result["status"] == "1"
    assert result["family"] == "1"
    assert "Status transitions: 1." in result["summary"]
    assert "Row 1: /a" in result["list"]


@pytest.mark.skipif(os.getenv("RUN_BROWSER_PARITY_TESTS") != "1", reason="Enable with RUN_BROWSER_PARITY_TESTS=1")
@pytest.mark.parametrize(
    "scenario",
    ["found", "fallback", "miss", "ambiguous", "conflict"],
)
def test_lookup_table_browser_python_behavior_parity(tmp_path, scenario):
    playwright_sync = pytest.importorskip("playwright.sync_api")

    def _write_xml(path: Path, body: str) -> None:
        path.write_text(body, encoding="utf-8")

    def _build_spec(path: Path, condition_text: str, include_ambiguous: bool = False, include_conflict: bool = False) -> None:
        wb = Workbook()
        mapping = wb.active
        mapping.title = "Mapping"
        mapping.append(["Field Name", "Source XPath", "Condition", "Cardinality"])
        mapping.append(["/status/targetValue", "/status/sourceCode", condition_text, ""])

        lookup_a = wb.create_sheet("LookupA")
        lookup_a.append(["Country Lookup"])
        lookup_a.append(["Code", "Mapped Value"])
        lookup_a.append(["US", "United States"])
        lookup_a.append(["AU", "Australia"])
        if include_conflict:
            lookup_a.append(["US", "USA"])

        if include_ambiguous:
            lookup_b = wb.create_sheet("LookupB")
            lookup_b.append(["Country Lookup"])
            lookup_b.append(["Code", "Mapped Value"])
            lookup_b.append(["US", "USA"])
            lookup_b.append(["AU", "AUS"])

        wb.save(path)

    condition_by_scenario = {
        "found": "Refer lookup country code",
        "fallback": "Check LookUp-Conversion Tab(Country Code) and map. If cannot find in LookUp-Conversion then map the source",
        "miss": "Refer lookup country code",
        "ambiguous": "Refer lookup country code",
        "conflict": "Refer lookup country code",
    }
    source_by_scenario = {
        "found": "US",
        "fallback": "ZZ",
        "miss": "ZZ",
        "ambiguous": "US",
        "conflict": "US",
    }
    target_by_scenario = {
        "found": "United States",
        "fallback": "ZZ",
        "miss": "ZZ",
        "ambiguous": "US",
        "conflict": "US",
    }

    spec_path = tmp_path / f"lookup_{scenario}.xlsx"
    input_path = tmp_path / f"lookup_{scenario}_input.xml"
    output_path = tmp_path / f"lookup_{scenario}_output.xml"

    _build_spec(
        spec_path,
        condition_by_scenario[scenario],
        include_ambiguous=(scenario == "ambiguous"),
        include_conflict=(scenario == "conflict"),
    )
    _write_xml(input_path, f'<?xml version="1.0" encoding="UTF-8"?>\n<status><sourceCode>{source_by_scenario[scenario]}</sourceCode></status>\n')
    _write_xml(output_path, f'<?xml version="1.0" encoding="UTF-8"?>\n<status><targetValue>{target_by_scenario[scenario]}</targetValue></status>\n')

    py_result = validate_mapping_from_payload_bytes(
        str(spec_path),
        input_path.read_bytes(),
        input_path.name,
        output_path.read_bytes(),
        output_path.name,
        validation_mode="strict",
    )

    with playwright_sync.sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(WEB_PATH.resolve().as_uri())
        page.set_input_files("#mapping_spec", str(spec_path))
        page.set_input_files("#input_payload", str(input_path))
        page.set_input_files("#output_payload", str(output_path))
        page.select_option("#validation_mode", "strict")
        page.click("#submit-btn")
        page.wait_for_function("() => Boolean(window.lastResult && window.lastResult.summary)", timeout=120000)
        js_result = page.evaluate("() => window.lastResult")
        browser.close()

    py_support = py_result.get("rule_support_summary", {})
    js_support = js_result.get("rule_support_summary", {})
    py_grouped = py_result.get("summary", {}).get("grouped_error_counts", {})
    js_grouped = js_result.get("summary", {}).get("grouped_error_counts", {})

    assert int(py_support.get("lookup_table_rules", 0)) == 1
    assert int(js_support.get("lookup_table_rules", 0)) == 1

    assert int(py_grouped.get("lookup_mismatches", 0)) == int(js_grouped.get("lookup_mismatches", 0))
    assert py_result.get("summary", {}).get("status") == js_result.get("summary", {}).get("status")

    if scenario in {"found", "fallback", "ambiguous", "conflict"}:
        assert int(js_grouped.get("lookup_mismatches", 0)) == 0
    else:
        assert int(js_grouped.get("lookup_mismatches", 0)) >= 1
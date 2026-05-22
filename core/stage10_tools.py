from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from core.spec_reader import extract_rules, read_mapping_table


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def flatten_issue_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    diagnostics = result.get("error_diagnostics") or []
    if isinstance(diagnostics, list) and diagnostics:
        for item in diagnostics:
            if not isinstance(item, dict):
                continue
            details = item.get("details") if isinstance(item.get("details"), dict) else {}
            rows.append(
                {
                    "rule_row": item.get("rule_row", "-"),
                    "field_path": _clean(item.get("target_xpath") or item.get("source_xpath") or details.get("required_path") or details.get("actual")),
                    "issue_type": _clean(item.get("rule_stat") or item.get("issue_type") or "validation_issue"),
                    "description": _clean(item.get("message") or "Validation issue"),
                    "suggested_fix": _clean(item.get("suggested_fix") or "Review mapping rule and payload values."),
                }
            )

    if rows:
        return rows

    for error in result.get("errors", []) or []:
        text = _clean(error)
        row_match = re.search(r"Row\s+(\d+)", text, flags=re.IGNORECASE)
        target_match = re.search(r"\|\s*Target:\s*([^|]+)", text, flags=re.IGNORECASE)
        rows.append(
            {
                "rule_row": row_match.group(1) if row_match else "-",
                "field_path": _clean(target_match.group(1) if target_match else ""),
                "issue_type": "validation_error",
                "description": text,
                "suggested_fix": "Review this rule and update mapping or payload accordingly.",
            }
        )
    return rows


def write_excel_report(result: dict[str, Any], output_path: str) -> Path:
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    summary = result.get("summary") or {}
    human = result.get("human_summary") or {}
    rows = flatten_issue_rows(result)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Status", _clean(summary.get("status"))])
    ws_summary.append(["Error Count", int(result.get("error_count", 0) or 0)])
    ws_summary.append(["Validation Mode", _clean(result.get("validation_mode"))])
    ws_summary.append(["Headline", _clean(human.get("headline"))])

    for item in human.get("issue_breakdown", []) or []:
        if isinstance(item, dict):
            ws_summary.append([f"Issue: {_clean(item.get('issue'))}", int(item.get("count", 0) or 0)])

    ws_issues = wb.create_sheet("Issues")
    ws_issues.append(["Rule Row", "Field Path", "Issue Type", "Issue Description", "Suggested Fix"])
    for row in rows:
        ws_issues.append(
            [
                row.get("rule_row", "-"),
                row.get("field_path", ""),
                row.get("issue_type", "validation_issue"),
                row.get("description", ""),
                row.get("suggested_fix", ""),
            ]
        )

    ws_triage = wb.create_sheet("Triage")
    ws_triage.append(["Rule Row", "Issue", "Status", "Owner", "Notes"])
    for row in rows:
        ws_triage.append([row.get("rule_row", "-"), row.get("description", ""), "needs_fix", "", ""])

    wb.save(out_path)
    return out_path


def _rule_signature(rule: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        _clean(rule.get("target_xpath")),
        _clean(rule.get("source_xpath")),
        _clean(rule.get("cardinality")),
        _clean(rule.get("condition")),
        _clean(rule.get("note")),
    )


def _rule_identity(rule: dict[str, Any]) -> tuple[str, str]:
    return (_clean(rule.get("target_xpath")), _clean(rule.get("source_xpath")))


def diff_specs(base_spec_path: str, compare_spec_path: str) -> dict[str, Any]:
    base_df = read_mapping_table(base_spec_path)
    cmp_df = read_mapping_table(compare_spec_path)
    base_rules = extract_rules(base_df)
    cmp_rules = extract_rules(cmp_df)

    base_counter = Counter(_rule_signature(rule) for rule in base_rules)
    cmp_counter = Counter(_rule_signature(rule) for rule in cmp_rules)

    added_signatures = list((cmp_counter - base_counter).elements())
    removed_signatures = list((base_counter - cmp_counter).elements())

    base_by_identity: dict[tuple[str, str], dict[str, Any]] = {}
    for rule in base_rules:
        base_by_identity[_rule_identity(rule)] = rule

    cmp_by_identity: dict[tuple[str, str], dict[str, Any]] = {}
    for rule in cmp_rules:
        cmp_by_identity[_rule_identity(rule)] = rule

    changed: list[dict[str, Any]] = []
    shared_ids = set(base_by_identity.keys()) & set(cmp_by_identity.keys())
    for ident in sorted(shared_ids):
        old_rule = base_by_identity[ident]
        new_rule = cmp_by_identity[ident]
        if _rule_signature(old_rule) == _rule_signature(new_rule):
            continue
        changed.append(
            {
                "target_xpath": ident[0],
                "source_xpath": ident[1],
                "old": {
                    "cardinality": _clean(old_rule.get("cardinality")),
                    "condition": _clean(old_rule.get("condition")),
                    "note": _clean(old_rule.get("note")),
                },
                "new": {
                    "cardinality": _clean(new_rule.get("cardinality")),
                    "condition": _clean(new_rule.get("condition")),
                    "note": _clean(new_rule.get("note")),
                },
                "risk": "high" if _clean(old_rule.get("condition")) != _clean(new_rule.get("condition")) else "medium",
            }
        )

    def _sig_to_dict(sig: tuple[str, str, str, str, str]) -> dict[str, str]:
        return {
            "target_xpath": sig[0],
            "source_xpath": sig[1],
            "cardinality": sig[2],
            "condition": sig[3],
            "note": sig[4],
            "risk": "high" if sig[3] or sig[2] else "low",
        }

    added = [_sig_to_dict(sig) for sig in added_signatures]
    removed = [_sig_to_dict(sig) for sig in removed_signatures]

    total_delta = len(added) + len(removed) + len(changed)
    status = "PASS" if total_delta == 0 else "PASS_WITH_WARNINGS"

    return {
        "summary": {
            "status": status,
            "error_count": 0,
            "grouped_error_counts": {},
            "top_critical_errors": [],
            "parser_status": "clean",
            "parser_confidence": "high",
        },
        "human_summary": {
            "headline": f"Spec diff complete: {len(added)} added, {len(removed)} removed, {len(changed)} changed rule(s)",
            "issue_breakdown": [
                {"issue": "Rules added", "count": len(added)},
                {"issue": "Rules removed", "count": len(removed)},
                {"issue": "Rules changed", "count": len(changed)},
            ],
            "what_to_fix_first": [
                f"Review high-risk change at target '{item['target_xpath']}'"
                for item in changed
                if item.get("risk") == "high"
            ][:20],
            "checked_rules": len(base_rules) + len(cmp_rules),
            "skipped_rules": 0,
        },
        "valid": True,
        "validation_mode": "spec_diff",
        "strict_would_fail": False,
        "checked_rules": len(base_rules) + len(cmp_rules),
        "warnings": [],
        "rule_stats": {},
        "structure_summary": None,
        "semantic_summary": {},
        "rule_gap_summary": {},
        "mandatory_preflight": {},
        "reverse_validation_summary": {},
        "mapping_completeness": {},
        "structure_findings": [],
        "parser_diagnostics": {"status": "clean", "confidence": "high", "warnings": []},
        "rule_support_summary": {},
        "rule_decisions": [],
        "error_diagnostics": [],
        "skipped_rules": [],
        "unsupported_rule_suggestions": [],
        "error_sections": {},
        "top_critical_errors": [],
        "error_count": 0,
        "inputs": {"base_spec_path": base_spec_path, "compare_spec_path": compare_spec_path},
        "errors": [],
        "spec_diff": {
            "added": added,
            "removed": removed,
            "changed": changed,
            "summary": {
                "base_rules": len(base_rules),
                "compare_rules": len(cmp_rules),
                "added": len(added),
                "removed": len(removed),
                "changed": len(changed),
            },
        },
    }


def _normalize_path_token(token: str) -> str:
    token = _clean(token)
    if not token:
        return ""
    token = re.sub(r"\[.*?\]", "", token)
    token = token.replace("@", "")
    token = token.replace("-", "_")
    token = token.strip("/")
    if not token:
        return ""
    if re.match(r"^\d", token):
        token = f"f_{token}"
    return token


def _path_to_segments(path_text: str) -> list[str]:
    text = _clean(path_text)
    if not text:
        return []
    parts = re.split(r"[\n|]", text)
    primary = _clean(parts[0])
    if not primary:
        return []
    segments = []
    for raw in primary.split("/"):
        seg = _normalize_path_token(raw)
        if seg:
            segments.append(seg)
    return segments


def _set_nested_value(root: dict[str, Any], segments: list[str], value: Any) -> None:
    if not segments:
        return
    node = root
    for segment in segments[:-1]:
        if segment not in node or not isinstance(node[segment], dict):
            node[segment] = {}
        node = node[segment]
    node[segments[-1]] = value


def _placeholder_for_rule(rule: dict[str, Any], mode: str, index: int) -> str:
    card = _clean(rule.get("cardinality"))
    condition = _clean(rule.get("condition"))
    if mode == "template":
        return "<required>" if card in {"1..1", "1", "m"} else ""
    if mode == "sample":
        if condition:
            return f"sample_{index}_when_condition_met"
        return f"sample_{index}"
    return f"full_{index}"


def generate_payload_bundle(spec_path: str, generation_mode: str) -> dict[str, Any]:
    mode = _clean(generation_mode).lower()
    if mode not in {"template", "sample", "full"}:
        raise ValueError("generation_mode must be one of: template, sample, full")

    df = read_mapping_table(spec_path)
    rules = extract_rules(df)

    input_payload: dict[str, Any] = {}
    output_payload: dict[str, Any] = {}

    for idx, rule in enumerate(rules, start=1):
        source_segments = _path_to_segments(rule.get("source_xpath", ""))
        target_segments = _path_to_segments(rule.get("target_xpath", ""))
        value = _placeholder_for_rule(rule, mode, idx)
        if source_segments:
            _set_nested_value(input_payload, source_segments, value)
        if target_segments:
            _set_nested_value(output_payload, target_segments, value)

    headline = {
        "template": "Payload template generated",
        "sample": "Sample payload pair generated",
        "full": "Full-field payload pair generated",
    }[mode]

    return {
        "summary": {
            "status": "PASS",
            "error_count": 0,
            "grouped_error_counts": {},
            "top_critical_errors": [],
            "parser_status": "clean",
            "parser_confidence": "high",
        },
        "human_summary": {
            "headline": f"{headline}: {len(rules)} rule(s) projected",
            "issue_breakdown": [],
            "what_to_fix_first": [],
            "checked_rules": len(rules),
            "skipped_rules": 0,
        },
        "valid": True,
        "validation_mode": f"payload_{mode}",
        "strict_would_fail": False,
        "checked_rules": len(rules),
        "warnings": [],
        "rule_stats": {},
        "structure_summary": None,
        "semantic_summary": {},
        "rule_gap_summary": {},
        "mandatory_preflight": {},
        "reverse_validation_summary": {},
        "mapping_completeness": {},
        "structure_findings": [],
        "parser_diagnostics": {"status": "clean", "confidence": "high", "warnings": []},
        "rule_support_summary": {},
        "rule_decisions": [],
        "error_diagnostics": [],
        "skipped_rules": [],
        "unsupported_rule_suggestions": [],
        "error_sections": {},
        "top_critical_errors": [],
        "error_count": 0,
        "inputs": {"spec_path": spec_path, "generation_mode": mode},
        "errors": [],
        "generated_payloads": {
            "mode": mode,
            "input": input_payload,
            "output": output_payload,
        },
    }


def write_generated_payload_files(result: dict[str, Any], output_prefix: str) -> dict[str, Path]:
    generated = result.get("generated_payloads") or {}
    input_payload = generated.get("input") or {}
    output_payload = generated.get("output") or {}

    prefix_path = Path(output_prefix)
    prefix_path.parent.mkdir(parents=True, exist_ok=True)
    input_path = prefix_path.with_name(f"{prefix_path.name}_input.json")
    output_path = prefix_path.with_name(f"{prefix_path.name}_output.json")

    input_path.write_text(json.dumps(input_payload, indent=2), encoding="utf-8")
    output_path.write_text(json.dumps(output_payload, indent=2), encoding="utf-8")

    return {"input": input_path, "output": output_path}
